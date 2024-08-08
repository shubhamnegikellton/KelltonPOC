import ast
import base64
import io
import logging
import os
import re
from io import StringIO

import pandas as pd
import requests
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import easyocr
from io import BytesIO
import numpy as np

# Configure the logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def chat_completion(prompt_value, system_value=""):
    client = OpenAI()
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "system": f"{system_value}",
                "role": "user",
                "content": f"{prompt_value}",
            }
        ],
        stream=False,
        temperature=0,
    )

    return response


def get_header(df):
    # Step 1: Identify the row from the excel which is having the column header
    header_row_index = df.apply(lambda row: row != "", axis=1).sum(axis=1).idxmax()
    header_row = df.loc[header_row_index]
    header_row_values = header_row[header_row != ""].tolist()
    return header_row_values, header_row_index


def get_end_of_table(df, header_row):
    # Step 2: Identify the end of table
    df = df.iloc[header_row:]

    # Define a threshold for consecutive empty rows to confirm the end of the table
    threshold = 2
    empty_row_count = 0
    end_row = df.index[-1]  # Start with the last index of the DataFrame

    # Iterate to find the end of the table
    for index, row in df.iterrows():
        if row.isnull().all() or all(row.astype(str).str.strip() == ""):
            empty_row_count += 1
            if empty_row_count >= threshold:
                end_row = index - threshold + 1
                break
        else:
            empty_row_count = 0

    return end_row


def get_mapping(header, desired_columns, row_data):
    output = {
        "Desired ColumnA": "Header1",
        "Desired ColumnB": "Header2",
        "Desired ColumnC": "",
    }
    prompt = f"""
    header: {header}
    desired columns: {desired_columns}
    sample_raw_data: {row_data}

    Provide the dictionary with mappings where each desired column is associated with its corresponding header from the raw data. If a desired column cannot be mapped, ensure it remains empty.
    Note: Item column can contain different types of services, hardware, devices and software such as professional services, subscription, routers and cabling etc. Item can be a some material, item but not Id or number.
    if you are able to identify description in source column then make sure to map it with only description of desired column.

    Example Output: {output}
    Note: do not provide any other text except the json mentioned above not even ```json.
    """

    system_prompt = f"""
    You are a helpful assistant specializing in mapping data columns.
    """

    response = chat_completion(prompt, system_prompt)
    response = response.choices[0].message.content
    return response


def get_rest_data_map(rest_data_df, desired_columns, suppliers):
    output = {
        "Desired ColumnA": "value 1",
        "Desired ColumnB": "value 2",
        "Desired ColumnC": "",
    }
    prompt = f"""
    desired columns: {desired_columns}
    raw_data: {rest_data_df}
    suppliers: {suppliers}

    Provide the dictionary with mappings where desired column is associated with its raw data. If a desired column cannot be mapped, ensure it remains empty.
    Here, we can only map these desired columns: Date, Country, City, Supplier, Quote #, Currency.
    Please check suppliers if any supplier name found in it and map with Supplier otherwise check throughly for Supplier.
    Also, Supplier should not be SingTel or Singapore telecommunication or SINGAPORE TELECOM HONG KONG LIMITED and Dataformat should be like "14-Nov-2024"
    Do not update country or city from company adress or company name.
    
    Currency should be in ISO 4217 format (e.g., USD, EUR).
    Example Output: {output}
    Note: do not provide any other text except the json mentioned above not even ```json.
    """

    system_prompt = f"""
    You are a helpful assistant specializing in mapping data columns.
    """

    response = chat_completion(prompt, system_prompt)
    response = response.choices[0].message.content
    return response


def convert_str_to_dict(mapped_dict_str):
    # Convert string to dictionary if needed
    try:
        mapped_dict = ast.literal_eval(mapped_dict_str)
        if not isinstance(mapped_dict, dict):
            raise ValueError("mapped_dict is not a valid dictionary")
    except (ValueError, SyntaxError) as e:
        print(f"Error converting mapped_dict: {e}")
        mapped_dict = {}
    return mapped_dict


def apply_restriction_on_df(new_df):
    # drop rows which have total cost is 0
    new_df = new_df[new_df["Total Cost"] != 0]

    # If a blank row is found, drop all rows from that index onward
    drop_index = None
    count = 0
    for i, row in new_df.iterrows():
        if row["Item"].strip() == "" or row["Description"].strip() == "":
            count += 1
            if count >= 2:
                drop_index = i
                break
    if drop_index is not None:
        new_df = new_df.loc[:drop_index]

    # If item and description are same drop those rows.
    new_df = new_df.drop_duplicates(subset=["Item", "Description"], keep="first")

    # Drop rows where both 'Item' and 'Description' are blank
    new_df = new_df[
        ~(
            (new_df["Item"].astype(str).str.strip() == "")
            | (new_df["Description"].astype(str).str.strip() == "")
        )
    ]
    return new_df


def update_city_country(row, city_to_country, country_list):
    for col in row.index:
        if col not in ["Country", "City"]:
            cell_value = row[col]
            if pd.isna(cell_value):
                continue

            cell_value_parts = re.split(r"\s*[,.;]\s*", str(cell_value).strip())
            for part in cell_value_parts:
                part = part.strip()
                if part in city_to_country:
                    row["Country"] = city_to_country[part]["country"]
                    row["City"] = part
                    return row
                elif part in country_list:
                    row["Country"] = part
                    row["City"] = None
                    return row
    return row


def extract_numeric(value):
    """Extract the numeric part from a mixed value (number + string)."""
    if isinstance(value, str):
        match = re.search(r"\d+\.?\d*", value)
        if match:
            return float(match.group(0))
    elif isinstance(value, (int, float)):
        return value
    return pd.NA


def update_unit_cost(df):
    df["Total Cost"] = df["Total Cost"].apply(extract_numeric)
    df["QTY"] = df["QTY"].apply(extract_numeric)
    df["Total Cost"] = pd.to_numeric(df["Total Cost"], errors="coerce")
    df["QTY"] = pd.to_numeric(df["QTY"], errors="coerce")
    df["Unit Cost"] = df["Total Cost"] / df["QTY"]

    return df


def get_exchange_rate(from_currency, to_currency="USD"):
    if from_currency:
        url = (
            f"https://api.frankfurter.app/latest?from={from_currency}&to={to_currency}"
        )
        response = requests.get(url)
        data = response.json()
        if "error" not in data:
            return data["rates"][to_currency]
        else:
            print(f"Error fetching exchange rate: {data['error']}")
    return 1


def convert_to_usd(amount, from_currency):
    if from_currency == "USD":
        return round(amount, 2)
    exchange_rate = get_exchange_rate(from_currency)
    if exchange_rate:
        return round(amount * exchange_rate, 2)
    else:
        return None


def city_country_mapped_list():
    script_dir = os.path.dirname(__file__)
    file_path = os.path.join(script_dir, "../files/worldcities.csv")
    file_path = os.path.abspath(file_path)
    city_country_df = pd.read_csv(file_path)
    city_country_df_aggregated = (
        city_country_df.groupby("city_ascii").first().reset_index()
    )
    city_to_country = city_country_df_aggregated.set_index("city_ascii")[
        ["country"]
    ].to_dict("index")
    country_list = city_country_df_aggregated["country"].unique()

    return city_to_country, country_list


def get_response(df_str, desired_columns):
    prompt = f"""
    Follow below rules for respective columns and do the mapping accordingly :
    Supplier column:
      - Try to identify supplier name from email if not mentioned explicitly

    Item & description column:
      - Item column can contain different types of services such as professional services, testing, installation, labour, support, subscription and cabling etc.
      - Item column cannot be empty.

    Currency column:
      - Identify the currency based on recognizable prefixes or symbols such as USD, $, ₹, EUR, £, ¥, etc. For instance, "$100" should be identified as "USD."

    Hours column:
      - Identify hours based on keywords or patterns like 4Hour, 24X7, a range between hours or days or weeks etc.
      - Process the hour string and calculate in hour value, for example 24x7x1 is 168 hours & 8x5xNBD = 40 hours
      - Keep it blank if not found instead of "NaN".

    Unit Cost column:
      - If the unit cost is not directly provided, calculate it by dividing the total cost by the quantity.

    Total Cost column:
      - The Total Cost column name can be total amount, total cost, total price etc.
      - If currency is in USD then "Unit Cost (Direct)" and "Unit cost (USD)" will be the same.
      - If currency is different than USD then convert "Unit Cost (Direct)" to respective USD and put under "Unit cost (USD)" column.

    Note:
      - there could be services mentioned in multiple columns. you need to closely identity as a service and add those rows as additional data.
      - Return the output in a csv format and all values should be in quote.
      - Example output format:
            "Date","Item","Description","Country","City","Supplier","Quote #","Currency","Total Cost","QTY","Hours","Unit Cost","Unit Cost (USD)"
            "28-Dec-2024","Support","Holidays (USD)","Algeria","","ABC LTD","USD","1558.5","4","2","1000"
      - Provide only tabular records in csv format and don't include ```csv in response provide only the CSV string.

    Need to map the below data to the following columns: {desired_columns}. The first row of the table data should be treated as the header.

    {df_str}
    """

    system_prompt = f"""
    You are an advanced data processing assistant. Your primary function is to map the data based on the provided instructions.
    Take your time, carefully go through each instrucction, do background work and do not hallucinate or create data on your own.
    """
    response = chat_completion(prompt, system_prompt)
    response = response.choices[0].message.content
    response_text = (
        response.replace("\r\n", "\n").replace("```csv", "").replace("```", "")
    )
    try:
        mapped_df = pd.read_csv(StringIO(response_text), encoding="utf-8")
        return mapped_df
    except pd.errors.ParserError as e:
        return f"Error parsing CSV: {e}"
    return response


def get_mapping_data(header, desired_columns, row_data):
    output = {
        "Desired ColumnA": "Header1",
        "Desired ColumnB": "Header2",
        "Desired ColumnC": "",
    }
    prompt = f"""
    header: {header}
    desired columns: {desired_columns}
    sample_raw_data: {row_data}

    Provide the dictionary with mappings where each desired column is associated with its corresponding header from the raw data. If a desired column cannot be mapped, ensure it remains empty.
    Note: Item column can contain different types of services, hardware, devices and software such as professional services, subscription, routers and cabling etc. Item can be a some material, item but not Id or number.
    if you are able to identify description in source column then make sure to map it with only description of desired column.

    Example Output: {output}
    Note: do not provide any other text except the json mentioned above not even ```json.
    """

    system_prompt = f"""
    You are a helpful assistant specializing in mapping data columns.
    """

    response = chat_completion(prompt, system_prompt)
    response = response.choices[0].message.content
    return response


def has_valid_header(df, threshold=0.5):
    """
    Check if the DataFrame has a valid header by ensuring that more than a threshold
    proportion of columns have names.
    """
    named_columns = [
        col for col in df.columns if not col.startswith("Unnamed") and col.strip()
    ]
    return len(named_columns) / len(df.columns) > threshold


def contains_bom_or_missing_price(columns):
    """
    Check if any of the column names contain 'BOM' or if 'price' is missing in the headers.
    """
    bom_pattern = re.compile(r"\bBOM\b", re.IGNORECASE)
    price_pattern = re.compile(r"\bprice\b", re.IGNORECASE)

    # Check if 'BOM' is in any column name
    contains_bom = any(bom_pattern.search(str(col)) for col in columns)

    # Check if 'price' is missing from the column names
    missing_price = not any(price_pattern.search(str(col)) for col in columns)

    # Return True if either condition is met
    return contains_bom or missing_price


def apply_restriction_on_c_df(new_df):
    # drop rows which have total cost is 0
    new_df = new_df[
        (new_df["Total Cost"] != 0) & (new_df["Total Cost"].astype(str) != "")
    ]

    # If item and description are same drop those rows.
    # new_df = new_df.drop_duplicates(subset=['Item', 'Description'], keep='first')

    # Drop rows where both 'Item' and 'Description' are blank
    new_df = new_df[
        ~(
            (new_df["Item"].astype(str).str.strip() == "")
            & (new_df["Description"].astype(str).str.strip() == "")
        )
    ]
    return new_df


def update_row(row, quotation_df):
    match_key = row["Match Key"]
    if pd.notna(match_key):  # Ensure match_key is not NaN
        matching_row = quotation_df[quotation_df["BOM"] == match_key]
        if not matching_row.empty:
            row["Country"] = matching_row.iloc[0]["Country"]
            row["City"] = matching_row.iloc[0]["Address (TO BE DEPLOYED)"]
            row["Supplier"] = matching_row.iloc[0]["Solution"]
            row["Quote #"] = matching_row.iloc[0]["Customer Site ID/Name"]
    return row


def update_city_value(city_value):
    city_to_country, iso2_to_country = city_country_iso_mapped_list()
    # Split the city value
    cities = re.split(r"\s*[,.;]\s*", str(city_value).strip())
    # print(cities)
    for city in cities:
        if city in city_to_country:
            return city
    return None


def update_country_value(country_value):
    city_to_country, iso2_to_country = city_country_iso_mapped_list()
    country_value = str(country_value).strip()
    if country_value in iso2_to_country:
        return iso2_to_country[country_value]
    return country_value


def city_country_iso_mapped_list():
    script_dir = os.path.dirname(__file__)
    file_path = os.path.join(script_dir, "../files/worldcities.csv")
    file_path = os.path.abspath(file_path)
    city_country_df = pd.read_csv(file_path)
    city_country_df_aggregated = (
        city_country_df.groupby("city_ascii").first().reset_index()
    )
    city_to_country = city_country_df_aggregated.set_index("city_ascii")[
        ["country"]
    ].to_dict("index")
    iso2_to_country = (
        city_country_df_aggregated[["iso2", "country"]]
        .drop_duplicates()
        .set_index("iso2")["country"]
        .to_dict()
    )

    return city_to_country, iso2_to_country


def parse_final_answer(output: str) -> str:
    logger.info(f"Initial Query: {output}")
    # Regular expression to match and extract the SQL query between ```sql and ```
    pattern = r"SQLQuery:\s*```sql(.*?)```"
    match = re.search(pattern, output, re.DOTALL)

    if match:
        sql_query = match.group(1).strip()  # Extract the SQL query and strip any surrounding whitespace
        logger.info(f"Final extracted SQL Query: {sql_query}")
        return sql_query

    sql_query = output.split("SQLQuery: ")[1]
    logger.info(f"Final extracted SQL Query: {sql_query}")
    return sql_query


def log_output(output: str) -> str:
    logger.info(f"Output Results: {output}")
    return output


def get_images_from_uploaded_file(uploaded_file):
    try:
        images = []
        file_contents = uploaded_file.read()

        # Use 'with' to ensure the stream remains open and is properly closed after use
        with BytesIO(file_contents) as file_stream:
            file_name = uploaded_file.name
            file_extension = file_name.split('.')[-1].lower()

            if file_extension == 'xlsx':
                wb = load_workbook(file_stream, data_only=True)
                sheet_names = wb.sheetnames

                if not sheet_names:
                    print("No sheets found in the workbook.")
                    return []

                first_sheet = wb[sheet_names[0]]
                image_loader = SheetImageLoader(first_sheet)

                for row in first_sheet.iter_rows():
                    for cell in row:
                        if image_loader.image_in(cell.coordinate):
                            img = image_loader.get(cell.coordinate)
                            images.append(img)

        return images
    except Exception as e:
        print(f"Error: {e}")
        return []

def extract_text_from_image(image):
    reader = easyocr.Reader(['en'])
    image_np = np.array(image)
    result = reader.readtext(image_np, detail=0)
    return " ".join(result)
